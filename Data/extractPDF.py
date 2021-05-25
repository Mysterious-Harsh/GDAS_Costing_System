from PyQt5 import QtCore
from openpyxl import load_workbook

import sqlite3


def insertAcc():

	acc = open( "accessory.txt", 'rb' )
	accessoryID = 1

	for i in acc.readlines():
		s = i.decode()
		ACC = s.split()
		temp = ACC[ 1 ].split( '/' )
		if len( temp ) > 1:
			code = "GD {:04d}/{}".format( int( temp[ 0 ] ), temp[ 1 ] )
		else:
			code = "GD {:04d}/".format( int( temp[ 0 ] ) )
		unit = ACC.pop()
		name = " ".join( ACC[ 2 : ] )
		qty = 1

		print( " %s : %s : %s" % ( code, name, unit.upper() ) )
		DBHandler.execute(
		    """insert into accessories (accessory_id,date,code,name,unit,quantity,rate,currency) values (?,?,?,?,?,?,?,?)""",
		    (
		        accessoryID, QtCore.QDateTime.currentDateTime().toPyDateTime(), code, name,
		        unit.upper(), qty, 0, "UGX"
		        )
		    )
		DBHandler.commit()
		accessoryID += 1


def insertMat():

	workbook = load_workbook( filename="materials.xlsx" )
	material = workbook[ "Sheet3" ]
	row = material.max_row

	for i in range( 1, row + 1 ):
		materialID = i
		temp = str( material.cell( row=i, column=1 ).value )
		if '/' not in list( temp ):
			temp = temp + "/"
		code = "GDIL {}".format( temp )

		name = str( material.cell( row=i, column=2 ).value )
		unit = "Kg/Mtr"
		qty = "{:05.2f}/06.40".format( material.cell( row=i, column=3 ).value )
		thickness = float( material.cell( row=i, column=4 ).value )
		DBHandler.execute(
		    """insert into materials (material_id,date,code,name,unit,quantity,thickness) values (?,?,?,?,?,?,?)""",
		    (
		        materialID, QtCore.QDateTime.currentDateTime().toPyDateTime(), code, name, unit,
		        qty, thickness
		        )
		    )

		DBHandler.commit()

		print( materialID, code, name, unit, qty, thickness )


if __name__ == "__main__":
	DBHandler = sqlite3.connect(
	    "GDAS.data", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
	    )
	DBHandler.execute( 'pragma foreign_keys=ON' )
	insertAcc()
	insertMat()

	DBHandler.close()